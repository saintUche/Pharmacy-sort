from openpyxl import workbook, load_workbook 
from map import calculate_distance
from calculators import distance_score, salary_score, total_score, distance_score_to_distance

wb = load_workbook('Pharmacy-sort\excel-files\preference.xlsx')
ws = wb.active
mysheet = wb.create_sheet("PreferenceExtract_location")


#filters to be decided by user
max_distance = 20
min_salary = 20000
max_salary = 30000
start_location = ""
salary_weight = 0.6
distance_weight = 0.4

def get_postcode(program_title):

    allowed_characters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890"

    
    ends_with = "1234567890"

    for character in program_title:
        if character not in allowed_characters:
            program_title = program_title.replace(character, " ")

    messy_postcode = program_title.split()

    postcode = 0
    count_1_pos = 0 
    for i in messy_postcode:
            
            i_letter_code1 = ord(i[0])
            i_letter_code2 = ord(i[len(i)-1])
            count_2_pos = count_1_pos + 1 
            adjacent_str = messy_postcode[count_2_pos]
            j_letter_code1 = ord(adjacent_str[len(adjacent_str)-1])
            j_letter_code2 = ord(adjacent_str[0])

            if (i_letter_code1 <= 90 and i_letter_code1 >= 65) and (j_letter_code1 <= 90 and j_letter_code1>= 65) and (i_letter_code2 <= 57 and i_letter_code2 >= 48) and (j_letter_code2 <= 57 and j_letter_code2 >= 48):
                postcode = i + " " + adjacent_str
                return postcode
            else:
                count_1_pos = count_1_pos + 1
                if count_1_pos == len(messy_postcode) - 1:
                    count_1_pos = 0
                    for x in messy_postcode:
                        if (len(x) > 4) and (ord(x[-3]) <= 57 and ord(x[-3]) >= 48):
                            postcode = x
                            return postcode
                     
   
    if postcode == 0:
        return ("no valid postcode")





programs_order = []

for row in range (2, 2716):

    def get_score(pro_score):
        return pro_score.get("program score")
    
    def get_program(pro_title):
        return pro_title.get("program name")
    
    def get_salary(pro_salary):
        return pro_salary.get("program salary")

    def get_distance(pro_distance):
        return pro_distance.get("program distance")
    
    row_number = str(row)

    program_title = ws['D'+row_number].value
    program_salary = ws['O'+row_number].value

    goal_location = get_postcode(program_title)
    
    if goal_location == "no valid postcode":
        distance = "no postcode"

    else:
       curr_program_distance_score = distance_score(goal_location, start_location, max_distance)
   
       if curr_program_distance_score !=0:
           
           distance = distance_score_to_distance(curr_program_distance_score, max_distance)
           if distance > 0:
               
               curr_program_salary_score = salary_score(program_salary, max_salary, min_salary)
               
               if curr_program_salary_score > 0:
                   
                   program_total_score = total_score(curr_program_salary_score, curr_program_distance_score, salary_weight, distance_weight)

                   value = {"program name": program_title, "program salary": program_salary, "program distance" : distance, "program score" : program_total_score}
        

                   programs_order.append(value)

    

programs_order.sort(key=get_score)


mysheet['A1'].value = "program name"
mysheet['B1'].value = "salary"
mysheet['C1'].value = "distance"
mysheet['D1'].value = "score out of 100"

count = 0
for item in programs_order:

    
    program_title = ws['D'+row_number].value
    mysheet['A' + str(count + 2)].value = item.get("program name")
    mysheet['B' + str(count + 2)].value = item.get("program salary")
    mysheet['C' + str(count + 2)].value = item.get("program distance")
    mysheet['D' + str(count + 2)].value = item.get("program score")

    count = count + 1



wb.save('preference.xlsx')